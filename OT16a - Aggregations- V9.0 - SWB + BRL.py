import pandas as pd
from openpyxl import load_workbook
from pandas.io.formats import excel
import os
from tkinter.filedialog import askopenfilename
import sys
import datetime

#  OBJECTIVE OF THIS CODE
#  We import:
#   - company level data for the components for the cost data of the key datasets;
#   - the aggregated data from OT16 as marked in spreadsheet "PR24OT16_Copy_1.3".
#  It imports the 2025-2030 and "Constant" data for each company.
#  We sum across the various companies to create aggregated data.
#  We automatically generate an f_output file to update fountain.
#  We create related pivot table to support QA
#  We only want the aggregated values to be added to the F_output, not the data we import.


# Change the display options
pd.set_option('display.max_columns', 40)
pd.set_option('display.width', 2000)

# Define the spreadsheet to get the f_inputs from and send the f_outputs to
initial_directory = os.getcwd()  # Identify the current working directory, i.e. where code is saved
filename = askopenfilename(initialdir=initial_directory)  # Show an "Open" dialog box + return the path to selected file
print(filename)


# Get QA code 3 from the F_Inputs sheets - used later in F_Outputs code
wbook = load_workbook(filename)
wsheet = wbook['F_Inputs']

# Import the data from the file
df_PR24 = pd.read_excel(filename, sheet_name="F_Inputs", skiprows=1)
df_PR24 = df_PR24[4:]  # Omit the four additional rows under the header
df_PR24.reset_index(drop=True, inplace=True)  # Reset the index

df_APR = pd.read_excel(filename, sheet_name="F_Inputs_APR", skiprows=1)
df_APR = df_APR[4:]  # Omit the four additional rows under the header
df_APR.reset_index(drop=True, inplace=True)  # Reset the index
#df_APR.rename(columns={"2022-23": "2025-30"}, inplace=True)

df = pd.concat([df_PR24, df_APR])


# Duplicates must not be in the F_Input report else they get double counted.
# The code below provides protection in case they exist.
print("\n *** ANALYSIS OF DUPLICATES *** ")
print("\nThe duplicated rows in the imported data are as follows:\n\n", df[df.duplicated()])
df.drop_duplicates(inplace=True)
print("\nAfter dropping duplicates, the remaining duplicated rows are as follows:\n\n", df[df.duplicated()]) #The result should be empty

# Analyse original dataframe
print("\n\n *** ANALYSIS OF IMPORTED DATA *** ")
original_bons = df["Reference"].unique().tolist()
print("\nThe number of unique bons in the imported data is:", len(original_bons))

original_companies = df["Acronym"].unique().tolist()
print("\nThe number of 'companies' used in the imported dataframe is:", len(original_companies))
print("\nThe companies used in the imported data dataframe are:", sorted(original_companies))

# Drop aggregated data to avoid double counting
print("\nPlease note the dataset already included the following aggregated data: \n", df[df['Acronym'].isin(["ENG", "WAL", "IND", "WASC", "WOC"]).dropna()])
df = df[~df['Acronym'].isin(["ENG", "WAL", "IND", "WASC", "WOC"])]
print("\nPlease note the dataset NOW includes the following aggregated data: \n", df[df['Acronym'].isin(["ENG", "WAL", "IND", "WASC", "WOC"]).dropna()])

altered_companies = df["Acronym"].unique().tolist()
print("\nThe number of companies used in the altered dataframe is:", len(altered_companies))
print("\nThe companies used in the altered dataframe is:", sorted(altered_companies))

# CALCULATE AGGREGATED DATA
years = ["Constant", "2022-23", "2024-25", "2025-26", "2026-27", "2027-28", "2028-29", "2029-30", "2025-30"]
columns_for_groupby = ["Reference"] + years
print(columns_for_groupby)

# Create a dataframe to hold the QA codes
df_QA = pd.DataFrame({"Reference": ["PR24QA_PR24OT01_OUT1", "PR24QA_PR24OT01_OUT2", "PR24QA_PR24OT01_OUT3", "PR24QA_PR24OT01_OUT4"]})

# Calculate England totals
companies_ENG = ['ANH', 'NES', 'SVE', 'SWB', 'SRN', 'TMS', 'NWT', 'WSX', 'YKY',
                 'AFW', 'BRL', 'PRT', 'SEW', 'SSC', 'SES']
df_ENG = df[df['Acronym'].isin(companies_ENG)]
df_ENG = df_ENG[columns_for_groupby].groupby(by='Reference', as_index=False).sum(min_count=1) # If fewer than min_count non-NA values are present the result will be NA.
df_ENG = pd.concat([df_ENG, df_QA])
df_ENG['Acronym'] = 'ENG'

# Calculate Wales totals
companies_WAL = ['WSH', 'HDD']
df_WAL = df[df['Acronym'].isin(companies_WAL)]
df_WAL = df_WAL[columns_for_groupby].groupby(by='Reference', as_index=False).sum(min_count=1)
df_WAL = pd.concat([df_WAL, df_QA])
df_WAL['Acronym'] = 'WAL'

# Calculate Sector totals
companies_IND = companies_ENG + companies_WAL
df_IND = df[df['Acronym'].isin(companies_IND)]
df_IND = df_IND[columns_for_groupby].groupby(by='Reference', as_index=False).sum(min_count=1)
df_IND = pd.concat([df_IND, df_QA])
df_IND['Acronym'] = 'IND'

# Calculate WASCs totals
companies_WASC = ['ANH','HDD', 'NES', 'SVE', 'SWB', 'SRN', 'TMS', 'NWT', 'WSX', 'WSH', 'YKY']
df_WASC = df[df['Acronym'].isin(companies_WASC)]
df_WASC = df_WASC[columns_for_groupby].groupby(by='Reference', as_index=False).sum(min_count=1)
df_WASC = pd.concat([df_WASC, df_QA])
df_WASC['Acronym'] = 'WASC'

# Calculate WOCs totals
companies_WOC = ['AFW', 'BRL', 'PRT', 'SEW', 'SSC', 'SES']
df_WOC = df[df['Acronym'].isin(companies_WOC)]
df_WOC = df_WOC[columns_for_groupby].groupby(by='Reference', as_index=False).sum(min_count=1)
df_WOC = pd.concat([df_WOC, df_QA])
df_WOC['Acronym'] = 'WOC'

# Create F_outputs
F_Outputs = pd.concat([df_ENG, df_IND, df_WAL, df_WASC, df_WOC])  # Add the three dataframe together
F_Outputs.loc[F_Outputs["Reference"] == "PR24QA_PR24OT01_OUT1", years] = "[...]" + str(datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
F_Outputs.loc[F_Outputs["Reference"] == "PR24QA_PR24OT01_OUT2", years] = filename
F_Outputs.loc[F_Outputs["Reference"] == "PR24QA_PR24OT01_OUT3", years] = wsheet['E1'].value
F_Outputs.loc[F_Outputs["Reference"] == "PR24QA_PR24OT01_OUT4", years] = 0
F_Outputs['Model'] = 'Price Review 2024'
F_Outputs['Item description'] = ''
F_Outputs['Unit'] = ''
F_Outputs = F_Outputs[['Acronym', 'Reference', 'Item description', 'Unit', 'Model'] + years]
F_Outputs.reset_index(inplace=True, drop=True)

# Pivot Table Creation to support QA
pivot_constant = pd.pivot_table(df, values="Constant", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_constant['ENG'] = pivot_constant[companies_ENG].sum(axis=1)
pivot_constant['WAL'] = pivot_constant[companies_WAL].sum(axis=1)
pivot_constant['WASC'] = pivot_constant[companies_WASC].sum(axis=1)
pivot_constant['WOC'] = pivot_constant[companies_WOC].sum(axis=1)
pivot_constant['IND'] = pivot_constant[companies_ENG + companies_WAL].sum(axis=1)

pivot_202223 = pd.pivot_table(df, values="2022-23", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202223['ENG'] = pivot_202223[companies_ENG].sum(axis=1)
pivot_202223['WAL'] = pivot_202223[companies_WAL].sum(axis=1)
pivot_202223['WASC'] = pivot_202223[companies_WASC].sum(axis=1)
pivot_202223['WOC'] = pivot_202223[companies_WOC].sum(axis=1)
pivot_202223['IND'] = pivot_202223[companies_ENG + companies_WAL].sum(axis=1)

pivot_202425 = pd.pivot_table(df, values="2024-25", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202425['ENG'] = pivot_202425[companies_ENG].sum(axis=1)
pivot_202425['WAL'] = pivot_202425[companies_WAL].sum(axis=1)
pivot_202425['WASC'] = pivot_202425[companies_WASC].sum(axis=1)
pivot_202425['WOC'] = pivot_202425[companies_WOC].sum(axis=1)
pivot_202425['IND'] = pivot_202425[companies_ENG + companies_WAL].sum(axis=1)

pivot_202526 = pd.pivot_table(df, values="2025-26", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202526['ENG'] = pivot_202526[companies_ENG].sum(axis=1)
pivot_202526['WAL'] = pivot_202526[companies_WAL].sum(axis=1)
pivot_202526['WASC'] = pivot_202526[companies_WASC].sum(axis=1)
pivot_202526['WOC'] = pivot_202526[companies_WOC].sum(axis=1)
pivot_202526['IND'] = pivot_202526[companies_ENG + companies_WAL].sum(axis=1)

pivot_202627 = pd.pivot_table(df, values="2026-27", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202627['ENG'] = pivot_202627[companies_ENG].sum(axis=1)
pivot_202627['WAL'] = pivot_202627[companies_WAL].sum(axis=1)
pivot_202627['WASC'] = pivot_202627[companies_WASC].sum(axis=1)
pivot_202627['WOC'] = pivot_202627[companies_WOC].sum(axis=1)
pivot_202627['IND'] = pivot_202627[companies_ENG + companies_WAL].sum(axis=1)

pivot_202728 = pd.pivot_table(df, values="2027-28", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202728['ENG'] = pivot_202728[companies_ENG].sum(axis=1)
pivot_202728['WAL'] = pivot_202728[companies_WAL].sum(axis=1)
pivot_202728['WASC'] = pivot_202728[companies_WASC].sum(axis=1)
pivot_202728['WOC'] = pivot_202728[companies_WOC].sum(axis=1)
pivot_202728['IND'] = pivot_202728[companies_ENG + companies_WAL].sum(axis=1)

pivot_202829 = pd.pivot_table(df, values="2028-29", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202829['ENG'] = pivot_202829[companies_ENG].sum(axis=1)
pivot_202829['WAL'] = pivot_202829[companies_WAL].sum(axis=1)
pivot_202829['WASC'] = pivot_202829[companies_WASC].sum(axis=1)
pivot_202829['WOC'] = pivot_202829[companies_WOC].sum(axis=1)
pivot_202829['IND'] = pivot_202829[companies_ENG + companies_WAL].sum(axis=1)

pivot_202930 = pd.pivot_table(df, values="2029-30", index="Reference", columns="Acronym", dropna=False) # NB This will keep
pivot_202930['ENG'] = pivot_202930[companies_ENG].sum(axis=1)
pivot_202930['WAL'] = pivot_202930[companies_WAL].sum(axis=1)
pivot_202930['WASC'] = pivot_202930[companies_WASC].sum(axis=1)
pivot_202930['WOC'] = pivot_202930[companies_WOC].sum(axis=1)
pivot_202930['IND'] = pivot_202930[companies_ENG + companies_WAL].sum(axis=1)

pivot_202530 = pd.pivot_table(df, values="2025-30", index="Reference", columns="Acronym",dropna=False) # NB This will keep
pivot_202530['ENG'] = pivot_202530[companies_ENG].sum(axis=1)
pivot_202530['WAL'] = pivot_202530[companies_WAL].sum(axis=1)
pivot_202530['WASC'] = pivot_202530[companies_WASC].sum(axis=1)
pivot_202530['WOC'] = pivot_202530[companies_WOC].sum(axis=1)
pivot_202530['IND'] = pivot_202530[companies_ENG + companies_WAL].sum(axis=1)

# Export results
excel.ExcelFormatter.header_style = None  # Get rid of default formatting
F_Outputs.set_index('Acronym', inplace=True)  # Drop the index
# Append the f_output sheet to the existing file, replacing any existing f_output sheet
with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists='replace') as writer:
    pivot_constant.to_excel(writer, sheet_name="Pivot Constant")
    pivot_202223.to_excel(writer, sheet_name="Pivot 2022-23")
    pivot_202425.to_excel(writer, sheet_name="Pivot 2024-25")
    pivot_202526.to_excel(writer, sheet_name="Pivot 2025-26")
    pivot_202627.to_excel(writer, sheet_name="Pivot 2026-27")
    pivot_202728.to_excel(writer, sheet_name="Pivot 2027-28")
    pivot_202829.to_excel(writer, sheet_name="Pivot 2028-29")
    pivot_202930.to_excel(writer, sheet_name="Pivot 2029-30")
    pivot_202530.to_excel(writer, sheet_name="Pivot 2025-30")
    F_Outputs.to_excel(writer, sheet_name="F_Outputs", startrow=1)

# Changes to the Excel file
wbook = load_workbook(filename)
wsheet = wbook['F_Outputs']
code_path = sys.argv[0]  # This gets the full path of this python script
wsheet['B1'] =  code_path
wsheet['C1'] = "F_Output was created on " + str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))  # Add in some description of the model
wsheet = wsheet.insert_rows(3,1)  # Add in an addtional row below the header row
wbook.save(filename)

# END
